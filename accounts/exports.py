"""
Utilitaires d'export PDF et Excel.
Utilisé par tous les modules pour générer des rapports.
"""

import io
from datetime import datetime
from django.http import HttpResponse

# PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ============================================================
# COULEURS DU THÈME
# ============================================================
VERT_PRIMAIRE = colors.HexColor("#27AE60")
VERT_FONCE = colors.HexColor("#1E8449")
BLANC = colors.white
GRIS_CLAIR = colors.HexColor("#F2F2F2")
NOIR = colors.black


# ============================================================
# EXPORT PDF
# ============================================================

def exporter_pdf(titre, colonnes, donnees, nom_fichier="rapport"):
    """
    Génère un fichier PDF avec un tableau de données.
    Args:
        titre: Titre du rapport
        colonnes: Liste des en-têtes de colonnes
        donnees: Liste de listes (lignes de données)
        nom_fichier: Nom du fichier sans extension
    Returns:
        HttpResponse avec le PDF
    """
    buffer = io.BytesIO()
    orientation = landscape(A4) if len(colonnes) > 5 else A4
    doc = SimpleDocTemplate(buffer, pagesize=orientation, topMargin=1.5 * cm)

    elements = []
    styles = getSampleStyleSheet()

    # Titre
    style_titre = ParagraphStyle(
        "TitreRapport",
        parent=styles["Title"],
        fontSize=16,
        textColor=VERT_FONCE,
        spaceAfter=10,
    )
    elements.append(Paragraph(titre, style_titre))

    # Sous-titre avec date
    style_sous_titre = ParagraphStyle(
        "SousTitre",
        parent=styles["Normal"],
        fontSize=10,
        textColor=colors.gray,
    )
    date_str = datetime.now().strftime("%d/%m/%Y à %H:%M")
    elements.append(Paragraph(f"Généré le {date_str}", style_sous_titre))
    elements.append(Spacer(1, 0.5 * cm))

    # Tableau
    table_data = [colonnes] + donnees
    table = Table(table_data, repeatRows=1)

    style_tableau = TableStyle([
        # En-tête
        ("BACKGROUND", (0, 0), (-1, 0), VERT_PRIMAIRE),
        ("TEXTCOLOR", (0, 0), (-1, 0), BLANC),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        # Corps
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 1), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # Alternance de couleurs
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [BLANC, GRIS_CLAIR]),
        # Bordures
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("LINEBELOW", (0, 0), (-1, 0), 1.5, VERT_FONCE),
        # Padding
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ])
    table.setStyle(style_tableau)
    elements.append(table)

    # Pied de page
    elements.append(Spacer(1, 1 * cm))
    elements.append(Paragraph(
        f"UIST-2ITS — Total: {len(donnees)} enregistrement(s)",
        styles["Normal"],
    ))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nom_fichier}.pdf"'
    return response


# ============================================================
# EXPORT EXCEL
# ============================================================

def exporter_excel(titre, colonnes, donnees, nom_fichier="rapport"):
    """
    Génère un fichier Excel avec un tableau de données.
    Args:
        titre: Titre du rapport
        colonnes: Liste des en-têtes de colonnes
        donnees: Liste de listes (lignes de données)
        nom_fichier: Nom du fichier sans extension
    Returns:
        HttpResponse avec le fichier Excel
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titre[:31]  # Limite Excel pour le nom d'onglet

    # Styles
    vert_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    blanc_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    normal_font = Font(name="Calibri", size=10)
    gris_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    bordure = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    centre = Alignment(horizontal="center", vertical="center")

    # Titre du rapport
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(colonnes))
    titre_cell = ws.cell(row=1, column=1, value=titre)
    titre_cell.font = Font(name="Calibri", bold=True, size=14, color="1E8449")
    titre_cell.alignment = centre

    # Date
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(colonnes))
    date_cell = ws.cell(
        row=2, column=1,
        value=f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    )
    date_cell.font = Font(name="Calibri", size=9, color="888888")
    date_cell.alignment = centre

    # En-têtes (ligne 4)
    for col_idx, col_name in enumerate(colonnes, 1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.fill = vert_fill
        cell.font = blanc_font
        cell.alignment = centre
        cell.border = bordure

    # Données
    for row_idx, row_data in enumerate(donnees, 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = normal_font
            cell.border = bordure
            cell.alignment = Alignment(vertical="center")
            if (row_idx - 5) % 2 == 1:
                cell.fill = gris_fill

    # Ajuster la largeur des colonnes
    for col_idx, col_name in enumerate(colonnes, 1):
        max_len = len(str(col_name))
        for row_data in donnees:
            if col_idx - 1 < len(row_data):
                max_len = max(max_len, len(str(row_data[col_idx - 1])))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nom_fichier}.xlsx"'
    return response


def importer_excel_notes(fichier):
    """
    Importe des notes depuis un fichier Excel.
    Colonnes attendues: Matricule, Code_Cours, Note, Type_Evaluation
    Returns: dict avec 'importees', 'erreurs'
    """
    wb = openpyxl.load_workbook(fichier)
    ws = wb.active

    resultats = {"importees": 0, "erreurs": []}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]:
            continue

        try:
            matricule = str(row[0]).strip()
            code_cours = str(row[1]).strip()
            valeur_note = float(row[2])
            type_eval = str(row[3]).strip() if len(row) > 3 and row[3] else "Examen"

            if not (0 <= valeur_note <= 20):
                resultats["erreurs"].append(
                    f"Ligne {row_idx}: Note {valeur_note} hors limites (0-20)"
                )
                continue

            from etudiant.models import Etudiant
            from gestion1.models import Cours

            etudiant = Etudiant.objects.filter(id_user__matricule=matricule).first()
            if not etudiant:
                resultats["erreurs"].append(
                    f"Ligne {row_idx}: Matricule '{matricule}' introuvable"
                )
                continue

            cours = Cours.objects.filter(code_cours=code_cours).first()
            if not cours:
                resultats["erreurs"].append(
                    f"Ligne {row_idx}: Code cours '{code_cours}' introuvable"
                )
                continue

            from gestion2.models import Note
            Note.objects.create(
                id_etudiant=etudiant,
                id_cours=cours,
                valeur_note=valeur_note,
                type_evaluation=type_eval,
                statut_validation="En attente",
            )
            resultats["importees"] += 1

        except (ValueError, IndexError) as e:
            resultats["erreurs"].append(f"Ligne {row_idx}: {str(e)}")

    return resultats
